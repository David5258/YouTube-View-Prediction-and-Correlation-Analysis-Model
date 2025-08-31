from googleapiclient.discovery import build
from openpyxl import load_workbook
import isodate

API_KEY = 'Replace with your actual API key'

# Initialize the YouTube API client
youtube = build('youtube', 'v3', developerKey=API_KEY)

def get_video_stats(video_id):
    try:
        request = youtube.videos().list(
            part="snippet, contentDetails, statistics",
            id=video_id
        )
        response = request.execute()
        
        print(f"API Response: {response}")
        
        stats = response['items'][0]['statistics']
        snippet = response['items'][0]['snippet']
        content_details = response['items'][0]['contentDetails']

        # Get category ID and video length
        category_id = snippet.get('categoryId', 'Unknown')
        duration_iso = content_details.get('duration', 'PT0S')
        subscriber_view_count = stats.get('viewCount', '0')

        duration = isodate.parse_duration(duration_iso)
        readable_duration = f"{duration.total_seconds() // 60:.0f}:{int(duration.total_seconds() % 60):02d}"

        return {
            'viewCount': stats.get('viewCount', '0'),
            'likeCount': stats.get('likeCount', '0'),
            'dislikeCount': stats.get('dislikeCount', '0'),
            'commentCount': stats.get('commentCount', '0'),
            'categoryId': category_id,
            'duration': duration,
            'subscriberViewCount': subscriber_view_count
        }
    
    except Exception as e:
        print(f"An error occurred: {e}")
        return {
            'viewCount': '0',
            'likeCount': '0',
            'dislikeCount': '0',
            'commentCount': '0'
            
        }


def update_excel(video_id, stats):
    print(f"Updating Excel for video_id: {video_id} with stats: {stats}")

    workbook = load_workbook('file location') #Replace with the Excel sheet you wish to update
    
    sheet = workbook.active
    
    row = None
    for i in range(2, sheet.max_row + 1):  # Assuming headers are in the first row
        if sheet.cell(i, 1).value == video_id:
            row = i
            break
    if row is None:
        row = sheet.max_row + 1
        sheet.cell(row, 1, video_id)
    print(f"Row to update: {row}")
    # Update stats in the Excel sheet
    sheet.cell(row, 2, stats['viewCount'])
    sheet.cell(row, 3, stats['likeCount'])
    sheet.cell(row, 4, stats['dislikeCount'])
    sheet.cell(row, 5, stats['commentCount'])
    sheet.cell(row, 6, stats['categoryId'])
    sheet.cell(row, 7, stats['duration'])
    sheet.cell(row, 8, stats['subscriberViewCount'])

    #Can add more stats if you want to analze other metrics

    workbook.save('file location') #Replace with the Excel sheet you wish to update
    print("Excel updated and saved.")

def main():
    video_ids = ['List of video ID','etc']  # Replace with your actual YouTube video ID
    for video_id in video_ids:
        stats = get_video_stats(video_id)
        print(f"Retrieved stats: {stats}")
        update_excel(video_id, stats)

if __name__ == '__main__':
    main()
